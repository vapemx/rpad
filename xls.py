from openpyxl import Workbook
from os import remove, listdir
import re

def createxls(name, age, bornPlace, socialMedia, podiums, victories, years, nextevents, dates, temp, news):
    ls = listdir()
    for i in ls:
        match = re.findall("(\w+.xlsx)", i)
        if match:
            remove(i)


    wb = Workbook()
    ws = wb.active

    ws["A1"].value = "Nombre: "
    ws["B1"].value = name
    ws["A2"].value = "Edad: "
    ws["B2"].value = age
    ws["A3"].value = "Lugar de nacimiento: "
    ws["B3"].value = bornPlace
    ws["A4"].value = "Redes sociales: "
    ws["B4"].value = socialMedia[0]
    ws["B5"].value = socialMedia[1]
    ws["B6"].value = socialMedia[2]
    ws["A7"].value = "Podios: "
    ws["B7"].value = podiums
    ws["A8"].value = "Victorias: "
    ws["B8"].value = victories
    ws["A9"].value = "Años en F1: "
    ws["B9"].value = years

    ws["A11"].value = "Proximos eventos"
    ws["A12"].value = nextevents[0]
    ws["A13"].value = nextevents[3]
    ws["A14"].value = nextevents[4]
    ws["B11"].value = "Temp. Min."
    ws["B12"].value = temp.get("d1")[1]
    ws["B13"].value = temp.get("d2")[1]
    ws["B14"].value = temp.get("d3")[1]
    ws["C11"].value = "Temp. Max."
    ws["C12"].value = temp.get("d1")[0]
    ws["C13"].value = temp.get("d2")[0]
    ws["C14"].value = temp.get("d3")[0]
    ws["D11"].value = "Fecha."
    ws["D12"].value = dates[0]
    ws["D13"].value = dates[2]
    ws["D14"].value = dates[3]

    ws["E1"].value = "Noticias."
    ws["E2"].value = news[0]
    ws["E3"].value = news[1]
    ws["E4"].value = news[2]

    wb.save("info.xlsx")

    return "Archivo creado con éxito."